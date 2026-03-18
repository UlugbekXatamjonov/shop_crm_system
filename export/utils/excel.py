"""
============================================================
EXCEL HELPER — openpyxl asosida
============================================================
Funksiyalar:
  make_excel_response   — headers + rows → HttpResponse (.xlsx)
  make_template         — bo'sh shablon faylini yaratish
  parse_excel_upload    — yuklangan faylni o'qib dict list qaytarish
"""

import io

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


# ============================================================
# STIL KONSTANTALARI
# ============================================================

HEADER_FONT  = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL  = PatternFill(fill_type='solid', fgColor='2E75B6')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
DATA_ALIGN   = Alignment(vertical='center', wrap_text=True)


# ============================================================
# ASOSIY FUNKSIYALAR
# ============================================================

def make_excel_response(filename: str, headers: list[str], rows: list[list]) -> HttpResponse:
    """
    headers: ['Nomi', 'Miqdori', ...]
    rows:    [['Mahsulot A', 10, ...], ...]

    → HttpResponse (Content-Type: application/vnd.openxmlformats...)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ma\'lumotlar'
    ws.row_dimensions[1].height = 32

    # Header qator
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # Ma'lumot qatorlari
    for row_idx, row in enumerate(rows, start=2):
        ws.row_dimensions[row_idx].height = 20
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = DATA_ALIGN
            # Musbat/manfiy sonlarni yashil/qizil rang
            if isinstance(val, (int, float)) and col_idx > 1:
                cell.font = Font(
                    color='00B050' if val >= 0 else 'FF0000'
                )

    # Ustun kengliklarini avtomatik sozlash
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header))
        for row in rows:
            val = row[col_idx - 1] if col_idx - 1 < len(row) else ''
            max_len = max(max_len, len(str(val)) if val is not None else 0)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Freeze header
    ws.freeze_panes = 'A2'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def make_template(filename: str, headers: list[str], notes: dict[str, str] | None = None) -> HttpResponse:
    """
    Bo'sh shablon yaratadi. notes = {header: izoh_matni}
    Misol: {'kategoriya': 'Mavjud kategoriya nomini yozing'}
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Shablon'
    ws.row_dimensions[1].height = 36

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN

        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(len(header) + 6, 18)

        # Izoh qo'shish
        if notes and header in notes:
            from openpyxl.comments import Comment
            comment = Comment(notes[header], 'Shop CRM')
            comment.width  = 200
            comment.height = 80
            cell.comment = comment

    # 2-qator — bo'sh misol
    ws.row_dimensions[2].height = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def parse_excel_upload(file) -> list[dict]:
    """
    Yuklangan Excel faylni o'qib, dict list qaytaradi.
    1-qator — header (kalit), 2-qatordan boshlab — ma'lumot.

    Qaytaradi: [{'nom': 'Mahsulot A', 'sale_price': '5000', ...}, ...]
    """
    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    except Exception:
        raise ValueError("Fayl noto'g'ri formatda. Faqat .xlsx qabul qilinadi.")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError("Fayl bo'sh.")

    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    result  = []

    for row_data in rows[1:]:
        # Butunlay bo'sh qatorlarni o'tkazib yuborish
        if all(v is None or str(v).strip() == '' for v in row_data):
            continue
        row_dict = {}
        for h, v in zip(headers, row_data):
            row_dict[h] = str(v).strip() if v is not None else ''
        result.append(row_dict)

    wb.close()
    return result
